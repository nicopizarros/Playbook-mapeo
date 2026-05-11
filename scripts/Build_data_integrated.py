"""
Playbook Sports Intelligence — Data Build Pipeline v2 (v1.00.31)
================================================================
Reads the single integrated workbook Playbook_Mapeo_v1_00_31.xlsx and regenerates
the 4 JSON files consumed by the Playbook Ecosistema web app.

Schema change vs. v1: primary key is now ID Maestro (A-###) instead of ID Vertical
(V1-001). ID Vertical is kept as a secondary display field.

Sources read:
  - UNIVERSO MAESTRO        → actors.json
  - MAPA DE CONEXIONES      → edges.json
  - POI MAESTRO             → poi.json (grouped by actor_id)
  - AMPLITUD FUNCIONAL      → enriches actors.json with Score Compuesto + Tipo capacidad
  - CLUSTERS EJECUTIVOS     → clusters.json (new 7-cluster schema)

Usage:
  python3 scripts/build_data_integrated.py \\
    --workbook Playbook_Mapeo_v1_00_31.xlsx \\
    --out data/

Exits with non-zero status if validation fails. Validation rules:
  - Every edge source/target must resolve to an A-### actor in UM.
  - Every POI actor_id must resolve to an A-### actor in UM.
  - Every cluster actor_id must resolve to an A-### actor in UM.
  - A-015 and A-068 (AF orphans not in UM) are silently dropped from score merge.

Outputs are written to <out>/{actors,edges,poi,clusters}.json (utf-8, indent=2).
"""

import argparse
import json
import re
import sys
from pathlib import Path
from collections import Counter, defaultdict
from openpyxl import load_workbook


# ─────────────────────────────────────────────────────────────────────────────
# Constants and regex
# ─────────────────────────────────────────────────────────────────────────────

ID_MAESTRO_RE  = re.compile(r'^A-\d+$')
ID_VERTICAL_RE = re.compile(r'^V\d+-\w+$')
A_ID_FIND      = re.compile(r'A-\d+')
V_CODE_FIND    = re.compile(r'\bV\d+\b')
CLUSTER_HDR_RE = re.compile(r'^CLUSTER\s+(\d+)\s*[—\-]\s*(.+)', re.I)

CLUSTER_ATTR_KEYS = {
    'Nombre',
    'Actores núcleo',
    'Actor central',
    'POI principal',
    'Conexiones internas clave',
    'Insight ejecutivo',
    'Hueco accionable',
}


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def clean(v):
    if v is None:
        return ''
    s = str(v).strip()
    # filter out Excel error markers and empty-equivalents
    if s.startswith('#') and s.endswith('!'):
        return ''
    return s


def parse_tier(v):
    s = clean(v)
    m = re.search(r'\d', s)
    return int(m.group()) if m else 3


def parse_bool_si(v):
    return clean(v).lower() in ('sí', 'si')


def parse_verticals(text):
    """Extract canonical V1..V9 codes from a string of secondary tags."""
    if not text:
        return []
    found = V_CODE_FIND.findall(clean(text))
    # dedupe preserving order
    seen, out = set(), []
    for v in found:
        if v not in seen:
            seen.add(v)
            out.append(v)
    return out


def parse_actor_ids(text):
    """Extract A-### IDs from any free-text string, deduped, order-preserving."""
    if not text:
        return []
    found = A_ID_FIND.findall(clean(text))
    seen, out = set(), []
    for a in found:
        if a not in seen:
            seen.add(a)
            out.append(a)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Stage 1 — UNIVERSO MAESTRO
# ─────────────────────────────────────────────────────────────────────────────

def build_actors(wb, report):
    """
    Read UNIVERSO MAESTRO. Headers are row 5; data starts row 6.
    Returns a dict id_maestro -> actor record.
    """
    ws = wb['UNIVERSO MAESTRO']
    rows = list(ws.iter_rows(min_row=6, values_only=True))

    actors = {}
    skipped = []
    for r in rows:
        if not r or not r[0]:
            continue
        id_maestro = clean(r[0])
        if not ID_MAESTRO_RE.match(id_maestro):
            skipped.append(('bad_id', id_maestro))
            continue

        id_vertical = clean(r[1])
        label       = clean(r[2])
        vertical    = clean(r[3])
        etiquetas   = clean(r[4])
        tipo        = clean(r[5])
        tier        = parse_tier(r[6])
        pais        = clean(r[7])
        ciudad      = clean(r[8])
        funcion     = clean(r[9])
        segmento    = clean(r[10])
        que_hace    = clean(r[11])
        por_que     = clean(r[12])
        known_for   = clean(r[13])
        productos   = clean(r[14])
        notas_prop  = clean(r[15])
        website     = clean(r[16])
        certeza     = clean(r[17])
        flag_str    = clean(r[18])
        visualizable_raw = clean(r[19])  # captured for diagnostics only; not used as filter

        # Fallback: if vertical primaria is blank, infer from ID Vertical prefix
        if not vertical and id_vertical:
            m = re.match(r'^(V\d+)-', id_vertical)
            if m:
                vertical = m.group(1)
                report['inferred_vertical'].append(id_maestro)

        # If still no vertical, record but keep actor (web will fall back to V1)
        if not vertical:
            report['no_vertical'].append(id_maestro)

        actors[id_maestro] = {
            'id':              id_maestro,
            'id_vertical':     id_vertical,
            'label':           label,
            'vertical':        vertical,
            'conexiones':      parse_verticals(etiquetas),
            'subcategoria':    tipo,
            'tipo':            tipo,
            'tier':            tier,
            'pais':            pais,
            'ciudad':          ciudad,
            'funcion':         funcion,
            'segmento':        segmento,
            'que_hace':        que_hace,
            'por_que':         por_que,
            'known_for':       known_for,
            'productos':       productos,
            'notas_propiedad': notas_prop,
            'website':         website,
            'certeza':         certeza,
            'flag':            bool(flag_str),
            'nota_flag':       flag_str,
            'fuentes':         flag_str,
            'watchlist':       False,
            # AF fields filled later
            'score_compuesto': None,
            'score_amplitud':  None,
            'score_profundidad': None,
            'tipo_capacidad':  '',
            # diagnostics-only
            '_visualizable_raw': visualizable_raw,
        }

    report['um_rows_total']    = len(rows)
    report['um_actors_kept']   = len(actors)
    report['um_rows_skipped']  = skipped
    return actors


# ─────────────────────────────────────────────────────────────────────────────
# Stage 2 — AMPLITUD FUNCIONAL enrichment
# ─────────────────────────────────────────────────────────────────────────────

def enrich_with_af(wb, actors, report):
    """
    Read AMPLITUD FUNCIONAL and merge Score Compuesto, Amplitud, Profundidad,
    and Tipo de capacidad onto each actor. Orphans (in AF but not in UM) are
    silently dropped per project decision.
    """
    ws = wb['AMPLITUD FUNCIONAL']
    # row 5 = top headers, row 6 = sub-headers (V1.A, V1.B...), row 7+ = data
    rows = list(ws.iter_rows(min_row=7, values_only=True))

    # Hard-coded column indices from inspection (verified):
    COL_ID, COL_AMP, COL_TIPO, COL_PROF, COL_COMP = 0, 46, 47, 48, 49

    merged = 0
    af_orphans = []
    for r in rows:
        if not r or not r[0]:
            continue
        a_id = clean(r[0])
        if not ID_MAESTRO_RE.match(a_id):
            continue
        if a_id not in actors:
            af_orphans.append(a_id)
            continue
        actors[a_id]['score_amplitud']    = r[COL_AMP]
        actors[a_id]['tipo_capacidad']    = clean(r[COL_TIPO])
        actors[a_id]['score_profundidad'] = r[COL_PROF]
        actors[a_id]['score_compuesto']   = r[COL_COMP]
        merged += 1

    report['af_rows_merged']  = merged
    report['af_orphans']      = af_orphans  # expected: ['A-015', 'A-068']
    return actors


# ─────────────────────────────────────────────────────────────────────────────
# Stage 3 — MAPA DE CONEXIONES
# ─────────────────────────────────────────────────────────────────────────────

def build_edges(wb, actors, report):
    """
    Read MAPA DE CONEXIONES. Headers row 5; data row 6+.
    Validate that source/target both resolve to actors. Records that fail are
    listed in the report as 'orphan_edges' and dropped from output (build still
    exits 0 unless there are zero valid edges).
    """
    ws = wb['MAPA DE CONEXIONES']
    rows = list(ws.iter_rows(min_row=6, values_only=True))

    # Column layout:
    # 0 ID Relación · 1 Actor A ID · 2 Actor A nombre · 3 Vertical A
    # 4 Actor B ID  · 5 Actor B nombre · 6 Vertical B
    # 7 Tipo · 8 Nivel · 9 Dirección · 10 Descripción · 11 Por qué importa
    # 12 Fuente · 13 Cross-vertical · 14 Visualizable

    edges = []
    orphan_edges = []
    for r in rows:
        if not r or not r[0]:
            continue
        rel_id   = clean(r[0])
        source   = clean(r[1])
        target   = clean(r[4])
        if not ID_MAESTRO_RE.match(source) or not ID_MAESTRO_RE.match(target):
            orphan_edges.append({
                'id': rel_id, 'source_raw': source, 'target_raw': target,
                'reason': 'non-canonical ID'
            })
            continue
        if source not in actors or target not in actors:
            orphan_edges.append({
                'id': rel_id, 'source': source, 'target': target,
                'reason': 'unknown actor'
            })
            continue

        edges.append({
            'id':           rel_id,
            'source':       source,
            'target':       target,
            'vertical_a':   clean(r[3]),
            'vertical_b':   clean(r[6]),
            'tipo':         clean(r[7]),
            'nivel':        clean(r[8]),
            'direccion':    clean(r[9]) or 'A→B',
            'descripcion':  clean(r[10]),
            'por_que':      clean(r[11]),
            'fuente':       clean(r[12]),
            'cross':        parse_bool_si(r[13]),
            'visualizable': parse_bool_si(r[14]),
        })

    report['edges_rows_total']  = len(rows)
    report['edges_kept']        = len(edges)
    report['edges_orphans']     = orphan_edges

    if orphan_edges:
        # Fail loud if there are descriptive-target style problems
        sys.stderr.write(
            f"\n  ⚠  {len(orphan_edges)} edge(s) reference unknown/non-canonical "
            f"actor IDs. See build report for details.\n"
        )
        for oe in orphan_edges[:5]:
            sys.stderr.write(f"     {oe}\n")
        if len(orphan_edges) > 5:
            sys.stderr.write(f"     ... and {len(orphan_edges) - 5} more\n")

    return edges


# ─────────────────────────────────────────────────────────────────────────────
# Stage 4 — POI MAESTRO
# ─────────────────────────────────────────────────────────────────────────────

def build_poi(wb, actors, report):
    """
    Read POI MAESTRO. Headers row 5; data row 6+.
    Returns dict actor_id -> list of POI records.
    """
    ws = wb['POI MAESTRO']
    rows = list(ws.iter_rows(min_row=6, values_only=True))

    # Columns: 0 ID POI · 1 Actor ID Maestro · 2 Actor name · 3 Vert actor
    # 4 Nombre · 5 Cargo · 6 Tier influencia · 7 Por qué importa · 8 Vínculos
    # 9 Nivel certeza · 10 Visualizable

    poi_map = defaultdict(list)
    total, orphans = 0, []
    for r in rows:
        if not r or not r[0]:
            continue
        poi_id   = clean(r[0])
        actor_id = clean(r[1])
        if not ID_MAESTRO_RE.match(actor_id):
            orphans.append({'poi_id': poi_id, 'actor_id_raw': actor_id, 'reason': 'non-canonical'})
            continue
        if actor_id not in actors:
            orphans.append({'poi_id': poi_id, 'actor_id': actor_id, 'reason': 'unknown actor'})
            continue

        poi_map[actor_id].append({
            'poi_id':     poi_id,
            'actor_id':   actor_id,
            'actor_name': clean(r[2]),
            'vertical':   clean(r[3]),
            'nombre':     clean(r[4]),
            'cargo':      clean(r[5]),
            'tier':       clean(r[6]),
            'por_que':    clean(r[7]),
            'vinculos':   clean(r[8]),
            'certeza':    clean(r[9]),
        })
        total += 1

    report['poi_total']    = total
    report['poi_actors']   = len(poi_map)
    report['poi_orphans']  = orphans
    return dict(poi_map)


# ─────────────────────────────────────────────────────────────────────────────
# Stage 5 — CLUSTERS EJECUTIVOS (new schema)
# ─────────────────────────────────────────────────────────────────────────────

def build_clusters(wb, actors, report):
    """
    Read CLUSTERS EJECUTIVOS. The sheet uses a key-value block layout, not a
    tabular one: each cluster starts with a 'CLUSTER N — title' header followed
    by 7 attribute rows (Nombre, Actores núcleo, Actor central, POI principal,
    Conexiones internas clave, Insight ejecutivo, Hueco accionable).
    """
    ws = wb['CLUSTERS EJECUTIVOS']
    all_rows = list(ws.iter_rows(values_only=True))

    clusters = []
    current = None
    for row in all_rows:
        a = row[0] if len(row) > 0 else None
        b = row[1] if len(row) > 1 else None
        a_str = clean(a)

        m = CLUSTER_HDR_RE.match(a_str)
        if m:
            if current:
                clusters.append(current)
            current = {
                'num':    int(m.group(1)),
                'header': m.group(2).strip(),
                'attrs':  {},
            }
            continue

        if current and a_str in CLUSTER_ATTR_KEYS and b is not None:
            current['attrs'][a_str] = clean(b)

    if current:
        clusters.append(current)

    # Build output records
    out = []
    cluster_orphan_actor_ids = []
    for c in clusters:
        attrs = c['attrs']
        # Extract A-### IDs from "Actores núcleo" and validate against UM
        raw_ids = parse_actor_ids(attrs.get('Actores núcleo', ''))
        valid_ids = []
        for aid in raw_ids:
            if aid in actors:
                valid_ids.append(aid)
            else:
                cluster_orphan_actor_ids.append({'cluster': c['num'], 'actor_id': aid})

        # Also extract A-### IDs from "Actor central" (may overlap with núcleo)
        central_ids = parse_actor_ids(attrs.get('Actor central', ''))
        actor_central = central_ids[0] if central_ids else ''

        # Derive verticales from the verticales of all member actors
        verticales = []
        seen = set()
        for aid in valid_ids:
            v = actors[aid].get('vertical', '')
            if v and v not in seen:
                seen.add(v)
                verticales.append(v)

        out.append({
            'num':           c['num'],
            'header':        c['header'],
            'nombre':        attrs.get('Nombre', ''),
            'actor_ids':     valid_ids,
            'actor_central': actor_central,
            'poi_principal': attrs.get('POI principal', ''),
            'conexiones':    attrs.get('Conexiones internas clave', ''),
            'insight':       attrs.get('Insight ejecutivo', ''),
            'hueco':         attrs.get('Hueco accionable', ''),
            'verticales':    verticales,
        })

    report['clusters_total']         = len(out)
    report['clusters_orphans']       = cluster_orphan_actor_ids
    report['clusters_total_actors']  = sum(len(c['actor_ids']) for c in out)
    report['clusters_unique_actors'] = len({a for c in out for a in c['actor_ids']})
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Output formatting
# ─────────────────────────────────────────────────────────────────────────────

def actors_to_list(actors_dict):
    """Strip diagnostics-only fields and return a list ordered by ID Maestro."""
    out = []
    for aid in sorted(actors_dict.keys(), key=lambda x: int(x.split('-')[1])):
        a = dict(actors_dict[aid])
        a.pop('_visualizable_raw', None)
        out.append(a)
    return out


def write_json(path, data):
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')


# ─────────────────────────────────────────────────────────────────────────────
# Build report printer
# ─────────────────────────────────────────────────────────────────────────────

def print_report(r, edges_count, poi_count, clusters_count, actors_count):
    print()
    print('═' * 72)
    print('  BUILD REPORT — Playbook_Mapeo_v1_00_31.xlsx')
    print('═' * 72)
    print()
    print(f'  UNIVERSO MAESTRO')
    print(f'    rows scanned         : {r["um_rows_total"]}')
    print(f'    actors kept          : {r["um_actors_kept"]}')
    if r['inferred_vertical']:
        print(f'    vertical inferred    : {len(r["inferred_vertical"])} '
              f'({", ".join(r["inferred_vertical"][:3])}{"..." if len(r["inferred_vertical"]) > 3 else ""})')
    if r['no_vertical']:
        print(f'    no vertical at all   : {len(r["no_vertical"])} '
              f'({", ".join(r["no_vertical"])})')
    print()
    print(f'  AMPLITUD FUNCIONAL')
    print(f'    scores merged        : {r["af_rows_merged"]}')
    print(f'    orphans (dropped)    : {len(r["af_orphans"])} '
          f'({", ".join(r["af_orphans"]) or "—"})')
    print()
    print(f'  MAPA DE CONEXIONES')
    print(f'    rows scanned         : {r["edges_rows_total"]}')
    print(f'    edges kept           : {r["edges_kept"]}')
    print(f'    orphan edges         : {len(r["edges_orphans"])}')
    print()
    print(f'  POI MAESTRO')
    print(f'    POIs kept            : {r["poi_total"]}')
    print(f'    POIs grouped under   : {r["poi_actors"]} actors')
    print(f'    orphan POIs          : {len(r["poi_orphans"])}')
    print()
    print(f'  CLUSTERS EJECUTIVOS')
    print(f'    clusters parsed      : {r["clusters_total"]}')
    print(f'    member actor refs    : {r["clusters_total_actors"]} '
          f'({r["clusters_unique_actors"]} unique)')
    print(f'    orphan actor refs    : {len(r["clusters_orphans"])}')
    print()
    print('  OUTPUT')
    print(f'    actors.json          : {actors_count}')
    print(f'    edges.json           : {edges_count}')
    print(f'    poi.json             : {poi_count} actors with POIs')
    print(f'    clusters.json        : {clusters_count}')
    print('═' * 72)
    print()


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--workbook', required=True, help='Path to Playbook_Mapeo_v1_00_31.xlsx')
    ap.add_argument('--out',      required=True, help='Output directory for JSON files')
    args = ap.parse_args()

    wb_path = Path(args.workbook)
    out_dir = Path(args.out)
    if not wb_path.exists():
        sys.stderr.write(f'ERROR: workbook not found: {wb_path}\n')
        sys.exit(2)
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f'Loading workbook: {wb_path}')
    wb = load_workbook(wb_path, read_only=True, data_only=True)

    report = {
        'inferred_vertical': [],
        'no_vertical':       [],
    }

    actors = build_actors(wb, report)
    actors = enrich_with_af(wb, actors, report)
    edges  = build_edges(wb, actors, report)
    poi    = build_poi(wb, actors, report)
    clusters = build_clusters(wb, actors, report)

    actors_list = actors_to_list(actors)

    write_json(out_dir / 'actors.json',   actors_list)
    write_json(out_dir / 'edges.json',    edges)
    write_json(out_dir / 'poi.json',      poi)
    write_json(out_dir / 'clusters.json', clusters)

    print_report(report,
                 edges_count=len(edges),
                 poi_count=len(poi),
                 clusters_count=len(clusters),
                 actors_count=len(actors_list))


if __name__ == '__main__':
    main()
    
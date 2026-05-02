"""
Playbook Sports Intelligence — Data Build Script
================================================
Reads both source XLSXs and outputs three JSON files to /data/:
  - actors.json   ← from Mapeo_Tracker_Final_*.xlsx  (Universo Maestro)
  - edges.json    ← from Playbook_Mapeo_Conexiones_*.xlsx  (Mapa de Conexiones)
  - poi.json      ← from Mapeo_Tracker_Final_*.xlsx  (POI Maestro)
  - clusters.json ← from Playbook_Mapeo_Conexiones_*.xlsx  (Clusters de Poder)

Usage:
  python scripts/build_data.py \
    --tracker  "Mapeo_Tracker_Final_1_00.xlsx" \
    --conexiones "Playbook_Mapeo_Conexiones_Expandido_2026.xlsx"

Output is written to ../data/ relative to this script.
"""

import argparse
import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

# ── Helpers ─────────────────────────────────────────────────────────────────

ID_RE = re.compile(r'^(V\d+-\w+)')

def clean(val):
    """Strip whitespace from a cell value; return empty string for None."""
    if val is None:
        return ""
    return str(val).strip()

def parse_actor_id(raw):
    """Extract 'V1-001' style ID from a combined 'V1-001 Actor Name' string."""
    m = ID_RE.match(clean(raw))
    return m.group(1) if m else clean(raw)

def tier_num(tier_str):
    """Convert 'Tier 1' → 1, etc."""
    s = clean(tier_str)
    m = re.search(r'\d', s)
    return int(m.group()) if m else 3

def vertical_connections(etiquetas_str):
    """
    Parse secondary tags like 'V9 (Ollamani) · V7 (Estadio) · V2 (TUDN)'
    into a clean list of vertical codes: ['V9', 'V7', 'V2']
    """
    if not etiquetas_str:
        return []
    return list(dict.fromkeys(re.findall(r'\bV\d+\b', clean(etiquetas_str))))

# ── actors.json ──────────────────────────────────────────────────────────────

def build_actors(tracker_path):
    wb = load_workbook(tracker_path, read_only=True, data_only=True)
    ws = wb['UNIVERSO MAESTRO']
    rows = list(ws.iter_rows(min_row=4, values_only=True))

    actors = []
    for row in rows:
        if not row[0]:
            continue
        actor_id   = clean(row[0])
        name       = clean(row[1])
        vertical   = clean(row[2])
        etiquetas  = clean(row[3])
        subcategory= clean(row[4])
        tipo       = clean(row[5])
        visibilidad= clean(row[6])
        tier       = tier_num(row[7])
        pais       = clean(row[8])
        ciudad     = clean(row[9])
        que_hace   = clean(row[10])
        por_que    = clean(row[11])
        known_for  = clean(row[12])
        website    = clean(row[13])
        certeza    = clean(row[15])
        flag       = clean(row[16]).lower() == 'sí'
        nota_flag  = clean(row[17])
        watchlist  = clean(row[18]).lower() == 'sí'
        fuentes    = clean(row[21])

        actors.append({
            "id":          actor_id,
            "label":       name,
            "vertical":    vertical,
            "conexiones":  vertical_connections(etiquetas),
            "subcategoria":subcategory,
            "tipo":        tipo,
            "visibilidad": visibilidad,
            "tier":        tier,
            "pais":        pais,
            "ciudad":      ciudad,
            "que_hace":    que_hace,
            "por_que":     por_que,
            "known_for":   known_for,
            "website":     website,
            "certeza":     certeza,
            "flag":        flag,
            "nota_flag":   nota_flag,
            "watchlist":   watchlist,
            "fuentes":     fuentes,
        })

    print(f"  actors.json → {len(actors)} actors")
    return actors

# ── poi.json ─────────────────────────────────────────────────────────────────

def build_poi(tracker_path):
    wb = load_workbook(tracker_path, read_only=True, data_only=True)
    ws = wb['POI MAESTRO']
    rows = list(ws.iter_rows(min_row=4, values_only=True))

    # Group POIs by actor ID so each actor has a list of people
    poi_map = {}
    raw_list = []
    for row in rows:
        if not row[0]:
            continue
        poi_id     = clean(row[0])
        actor_id   = clean(row[1])
        actor_name = clean(row[2])
        vertical   = clean(row[3])
        nombre     = clean(row[4])
        cargo      = clean(row[5])
        por_que    = clean(row[6])
        conexion   = clean(row[7])
        certeza    = clean(row[8])
        linkedin   = clean(row[9])
        notas      = clean(row[11])

        entry = {
            "poi_id":     poi_id,
            "actor_id":   actor_id,
            "actor_name": actor_name,
            "vertical":   vertical,
            "nombre":     nombre,
            "cargo":      cargo,
            "por_que":    por_que,
            "conexion":   conexion,
            "certeza":    certeza,
            "linkedin":   linkedin,
            "notas":      notas,
        }
        raw_list.append(entry)
        if actor_id not in poi_map:
            poi_map[actor_id] = []
        poi_map[actor_id].append(entry)

    print(f"  poi.json → {len(raw_list)} POIs across {len(poi_map)} actors")
    return poi_map

# ── edges.json ───────────────────────────────────────────────────────────────

def build_edges(conexiones_path):
    wb = load_workbook(conexiones_path, read_only=True, data_only=True)
    ws = wb['MAPA DE CONEXIONES']
    rows = list(ws.iter_rows(min_row=3, values_only=True))

    edges = []
    for row in rows:
        if not row[0]:
            continue
        rel_id      = clean(row[0])
        source      = parse_actor_id(row[1])
        vertical_a  = clean(row[2])
        target      = parse_actor_id(row[3])
        vertical_b  = clean(row[4])
        tier_con    = clean(row[5])
        descripcion = clean(row[6])
        tipo        = clean(row[7])
        direccion   = clean(row[8])
        fuerza      = clean(row[9])
        fuente      = clean(row[10])
        vigente     = clean(row[11]).lower() == 'sí'
        cross       = clean(row[12]).lower() == 'sí'
        notas       = clean(row[13])

        edges.append({
            "id":          rel_id,
            "source":      source,
            "target":      target,
            "vertical_a":  vertical_a,
            "vertical_b":  vertical_b,
            "tier":        tier_con,
            "descripcion": descripcion,
            "tipo":        tipo,
            "direccion":   direccion,
            "fuerza":      fuerza,
            "fuente":      fuente,
            "vigente":     vigente,
            "cross":       cross,
            "notas":       notas,
        })

    print(f"  edges.json → {len(edges)} relationships")
    return edges

# ── clusters.json ────────────────────────────────────────────────────────────

def build_clusters(conexiones_path):
    wb = load_workbook(conexiones_path, read_only=True, data_only=True)
    ws = wb['CLUSTERS DE PODER']
    rows = list(ws.iter_rows(min_row=3, values_only=True))

    clusters = []
    for row in rows:
        if not row[0]:
            continue
        nombre      = clean(row[0])
        actores_raw = clean(row[1])
        verticales  = clean(row[2])
        puente      = clean(row[3])
        control     = clean(row[4])
        conexiones  = clean(row[5])

        # Extract actor IDs from the actores_raw string
        actor_ids = re.findall(r'V\d+-\w+', actores_raw)

        clusters.append({
            "nombre":     nombre,
            "actor_ids":  actor_ids,
            "verticales": verticales,
            "puente":     puente,
            "control":    control,
            "conexiones": conexiones,
        })

    print(f"  clusters.json → {len(clusters)} clusters")
    return clusters

# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Build Playbook data JSONs from XLSXs")
    parser.add_argument("--tracker",    required=True, help="Path to Mapeo_Tracker_Final_*.xlsx")
    parser.add_argument("--conexiones", required=True, help="Path to Playbook_Mapeo_Conexiones_*.xlsx")
    args = parser.parse_args()

    tracker_path    = Path(args.tracker)
    conexiones_path = Path(args.conexiones)
    out_dir         = Path(__file__).parent.parent / "data"
    out_dir.mkdir(exist_ok=True)

    for p in [tracker_path, conexiones_path]:
        if not p.exists():
            print(f"ERROR: File not found → {p}", file=sys.stderr)
            sys.exit(1)

    print("Building data files...")

    actors   = build_actors(tracker_path)
    poi      = build_poi(tracker_path)
    edges    = build_edges(conexiones_path)
    clusters = build_clusters(conexiones_path)

    (out_dir / "actors.json").write_text(
        json.dumps(actors, ensure_ascii=False, indent=2), encoding="utf-8")
    (out_dir / "poi.json").write_text(
        json.dumps(poi, ensure_ascii=False, indent=2), encoding="utf-8")
    (out_dir / "edges.json").write_text(
        json.dumps(edges, ensure_ascii=False, indent=2), encoding="utf-8")
    (out_dir / "clusters.json").write_text(
        json.dumps(clusters, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"\nDone. Files written to: {out_dir.resolve()}")

if __name__ == "__main__":
    main()
